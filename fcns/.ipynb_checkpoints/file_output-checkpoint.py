 
import os, time
from openpyxl import Workbook
from win32com.client import Dispatch

#column and worksheet definitions by index for ps0-ps3
col_layout =    {'0': {'1':'load','2':'vsense','3':'ripple','8':'dimon'},
                 '1': {'1':'load','2':'vsense','3':'ripple','8':'dimon'},
                 '2': {'1':'load','2':'vsense','3':'ripple','4':'vneg_ripple','5':'vpos_ripple','8':'dimon'},
                 '3': {'1':'load','2':'vsense','3':'ripple','4':'vneg_ripple','5':'vpos_ripple','8':'dimon'}}

worksheet_paste_origin = {'0':(11,16), 
                          '1':(41,16), 
                          '2':(70,16), 
                          '3':(99,16)}

def new_excel_file(filepath,dataset,pstate):
    ps=pstate #str(pstate)
    wb=Workbook(); ws1 = wb.create_sheet(title="loadline",index=1)
    file_name= os.path.join(filepath,'Loadline' + time.strftime("_%j_%M_%S", time.localtime()) +".xlsx")
    start_row=3; start_col=1
    for col,label in col_layout[ps].items(): 
            ws1.cell(column=int(col)+start_col-1, row=start_row, value=label)
    for row, data_dict in enumerate(dataset):
        for col,label in col_layout[ps].items():
                ws1.cell(column=int(col)+start_col-1, row=row+start_row+1, value=data_dict[label])
    wb.save(filename = file_name)

def intel_test_plan(filepathandname,dataset,pstate):
    ps=pstate #str(pstate)
    staticLLsheet='Static LL'
    start_row,start_col = worksheet_paste_origin[ps]
    excel = Dispatch("Excel.Application")
    excel.Visible = 1
    #excel.Workbooks.Cl
    wb = excel.Workbooks.Open(filepathandname)
    ws = wb.Sheets(staticLLsheet)
    ws.Select()
    for row, data_dict in enumerate(dataset):
        for col,label in col_layout[ps].items():
            if label in data_dict.keys():
                ws.Cells(row+start_row,int(col)+start_col-1).Value=data_dict[label]
    print("Results exported to Excel")
    #wb.Close(True)  #these last 2 lines were added because of suspicion excel files not being 
    #excel.Quit()    #close properly